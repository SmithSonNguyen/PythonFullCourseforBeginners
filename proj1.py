# Automation with Python
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1'] #Truy cập vào Sheet1
cell = sheet.cell(1, 1) #truy cập vào ô có địa chỉ (1, 1)

for row in range(2, sheet.max_row + 1): # [2, sheet.max_row + 1)
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values) #cung cấp data cho chart
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')
    
    